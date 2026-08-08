import streamlit as st
import pandas as pd
from datetime import date, datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client, Client

# Cấu hình trang cho di động
st.set_page_config(
    page_title="Chấm Lỗi",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# CSS tối ưu cho điện thoại
st.markdown("""
<style>
    /* Tối ưu kích thước cho mobile */
    .main {
        padding: 0.5rem;
    }
    .stButton > button {
        width: 100%;
        padding: 10px;
        font-size: 16px;
        font-weight: bold;
    }
    h1 {
        font-size: 24px;
        text-align: center;
    }
    h2 {
        font-size: 18px;
    }
    /* Nút tích lỗi sát lề trái */
    .error-button-container {
        width: 100%;
        margin-bottom: 8px;
    }
    .error-button {
        width: 100%;
        padding: 12px;
        border: 2px solid #ddd;
        border-radius: 5px;
        background-color: #f0f0f0;
        cursor: pointer;
        font-size: 15px;
        font-weight: 600;
        text-align: left;
        transition: all 0.2s ease;
        user-select: none;
    }
    .error-button:hover {
        border-color: #0066cc;
        background-color: #e6f0ff;
    }
    .error-button.active {
        background-color: #ff4444;
        color: white;
        border-color: #cc0000;
    }
    /* Định dạng bảng báo cáo */
    .total-row {
        font-weight: bold;
        background-color: #e8f4f8;
    }
</style>
""", unsafe_allow_html=True)

# Hàm phát âm thanh thông báo
def play_sound():
    """Phát âm thanh thông báo khi lưu thành công"""
    sound_html = """
    <audio autoplay>
        <source src="data:audio/wav;base64,UklGRiYAAABXQVZFZm10IBAAAAABAAEAQB8AAAB9AAACABAAZGF0YQIAAAAAAA==" type="audio/wav">
    </audio>
    """
    st.markdown(sound_html, unsafe_allow_html=True)

# ===== SUPABASE CONNECTION =====
@st.cache_resource
def get_supabase_client():
    """Kết nối tới Supabase"""
    try:
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_KEY"]
        supabase: Client = create_client(url, key)
        return supabase
    except Exception as e:
        st.error(f"❌ Lỗi kết nối Supabase: {e}")
        st.info("📌 Bạn cần setup `.streamlit/secrets.toml` với SUPABASE_URL và SUPABASE_KEY")
        return None

supabase = get_supabase_client()

# Danh sách cột lỗi
columns = [
    "Khong_that_day_an_toan", "Khong_bat_xi_nhan", "Khong_quan_sat_guong", 
    "Dung_do_xe_sai_quy_dinh", "Khong_chap_hanh_bien_bao", "Mo_cua_xe_khong_an_toan", 
    "Vuot_xe_khong_an_toan", "Quay_dau_sai_quy_dinh", "Khong_giam_toc_do", 
    "Khong_chap_hanh_vach_ke", "Khong_nghe_sat_hach_vien", "Loi_khac"
]

columns_display = [
    "Không thắt dây an toàn", "Không bật xi nhan trái/phải", "Không quan sát gương",
    "Dừng, đỗ xe sai quy định", "Không chấp hành hiệu lệnh", "Mở cửa xe không an toàn",
    "Vượt xe không đảm bảo", "Quay đầu xe không đúng", "Không quan sát, giảm tốc",
    "Lỗi vạch kẻ đường", "Không thực hiện theo yêu cầu", "Lỗi khác"
]

# Khởi tạo session state
if "selected_errors" not in st.session_state:
    st.session_state.selected_errors = set()

if "save_success" not in st.session_state:
    st.session_state.save_success = False

# ===== GIAO DIỆN NHẬP LIỆU =====
st.title("THỐNG KÊ LỖI ĐƯỜNG TRƯỜNG")

# Chọn ngày
ngay_sat_hach = st.date_input("📅 Ngày sát hạch", date.today(), format="DD/MM/YYYY")

# Form tích lỗi với nút bấm
st.write("**✓ Tích vào các lỗi học viên mắc phải (nhấn để chọn/bỏ chọn):**")

# Hiển thị thông báo nếu vừa lưu thành công
if st.session_state.save_success:
    st.success("✅ Đã lưu thành công!", icon="✅")
    st.balloons()
    st.session_state.save_success = False

# Tạo các nút bấm lỗi
for idx, display_text in enumerate(columns_display):
    is_selected = idx in st.session_state.selected_errors
    button_color = "🔴" if is_selected else "⚪"
    
    if st.button(f"{button_color} {display_text}", key=f"error_btn_{idx}", use_container_width=True):
        if idx in st.session_state.selected_errors:
            st.session_state.selected_errors.remove(idx)
        else:
            st.session_state.selected_errors.add(idx)
        st.rerun()

# Các nút hành động
st.divider()
col1, col2, col3 = st.columns(3)

with col1:
    if st.button("💾 Lưu", use_container_width=True):
        if not supabase:
            st.error("❌ Không thể kết nối Supabase")
        else:
            try:
                # Tạo object dữ liệu
                loi_values = [1 if idx in st.session_state.selected_errors else 0 for idx in range(len(columns))]
                
                data = {
                    "ngay": str(ngay_sat_hach),
                    **{columns[i]: loi_values[i] for i in range(len(columns))}
                }
                
                # Lưu vào Supabase
                response = supabase.table("loi_thi").insert([data]).execute()
                
                # Phát âm thanh
                play_sound()
                
                # Reset
                st.session_state.selected_errors = set()
                st.session_state.save_success = True
                st.rerun()
            except Exception as e:
                st.error(f"❌ Lỗi lưu dữ liệu: {e}")

with col2:
    if st.button("🔄 Xóa", use_container_width=True):
        st.session_state.selected_errors = set()
        st.rerun()

with col3:
    st.metric("Số lỗi chọn", len(st.session_state.selected_errors))

# ===== XUẤT BÁO CÁO =====
st.divider()
st.subheader("📊 Báo Cáo Tổng Hợp")

# Thống kê nhanh
if supabase:
    try:
        response = supabase.table("loi_thi").select("*").execute()
        df_all = pd.DataFrame(response.data) if response.data else pd.DataFrame()
        
        if not df_all.empty:
            st.info(f"📈 Tổng số lần ghi nhận: **{len(df_all)}** | Số ngày: **{df_all['ngay'].nunique()}**")
    except:
        pass

# Phần lọc dữ liệu
st.write("**Lọc dữ liệu báo cáo:**")
filter_col1, filter_col2, filter_col3 = st.columns(3)

filter_type = "all"
filter_date = None
filter_date_from = None
filter_date_to = None

with filter_col1:
    filter_type = st.radio("Chọn loại lọc:", ["Tất cả", "Một ngày", "Từ ngày đến ngày"], horizontal=True)

if filter_type == "Một ngày":
    with filter_col2:
        filter_date = st.date_input("Chọn ngày:", date.today(), format="DD/MM/YYYY", key="filter_single_date")
elif filter_type == "Từ ngày đến ngày":
    with filter_col2:
        filter_date_from = st.date_input("Từ ngày:", date.today() - timedelta(days=7), format="DD/MM/YYYY", key="filter_from_date")
    with filter_col3:
        filter_date_to = st.date_input("Đến ngày:", date.today(), format="DD/MM/YYYY", key="filter_to_date")

# Danh sách tên lỗi đầy đủ
full_error_names = [
    "Không thắt dây an toàn",
    "Không bật xi nhan trái/phải",
    "Không quan sát gương",
    "Dừng, đỗ xe sai quy định",
    "Không chấp hành hiệu lệnh",
    "Mở cửa xe không an toàn",
    "Vượt xe không đảm bảo",
    "Quay đầu xe không đúng",
    "Không quan sát, giảm tốc độ để hoạc dừng các trương hợp theo quy định",
    "Lỗi vạch kẻ đường",
    "Không thực hiện theo yêu cầu của Sát hạch viên",
    "Lỗi khác"
]

def create_report_excel(df_tong_hop):
    """Tạo file Excel với định dạng theo mẫu"""
    wb = Workbook()
    ws = wb.active
    ws.title = "TongHopLoi"
    
    # Định dạng các kiểu
    title_font = Font(name='Times New Roman', size=14, bold=True)
    total_font = Font(name='Times New Roman', size=11, bold=True, color="CC0000")
    header_font = Font(name='Times New Roman', size=10, bold=True)
    data_font = Font(name='Times New Roman', size=10)
    
    header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    
    # Hàng tiêu đề (Row 1)
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = "Số lượng lỗi do sát hạch viên trừ trong phần thi đường trường"
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    ws.row_dimensions[1].height = 30
    
    # Hàng tổng cộng (Row 2)
    total_row_data = df_tong_hop.iloc[0]
    ws['A2'] = "TỔNG CỘNG"
    ws['A2'].font = total_font
    ws['A2'].fill = total_fill
    ws['A2'].alignment = center_alignment
    ws['A2'].border = border
    
    col_idx = 2
    for col_name in df_tong_hop.columns[2:]:
        cell = ws.cell(row=2, column=col_idx)
        value = total_row_data[col_name]
        cell.value = int(value) if pd.notna(value) else 0
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center_alignment
        cell.border = border
        col_idx += 1
    
    # Hàng header (Row 3)
    col_idx = 1
    for col_name in df_tong_hop.columns:
        cell = ws.cell(row=3, column=col_idx)
        cell.value = col_name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
        col_idx += 1
    
    ws.row_dimensions[3].height = 40
    
    # Dữ liệu (từ Row 4 trở đi)
    for row_idx, (_, row_data) in enumerate(df_tong_hop.iloc[1:].iterrows(), start=4):
        col_idx = 1
        for col_name in df_tong_hop.columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row_data[col_name]
            
            if col_name == 'STT':
                cell.value = value if value != '' else ''
            elif col_name == 'Ngày sát hạch':
                cell.value = value
            else:
                cell.value = int(value) if pd.notna(value) else 0
            
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = border
            col_idx += 1
    
    # Điều chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    for col_idx in range(3, 15):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    
    # Lưu vào BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def format_date(date_str):
    """Chuyển đổi ngày sang định dạng DD/MM/YYYY"""
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        return date_obj.strftime("%d/%m/%Y")
    except:
        return date_str

if st.button("📥 Tạo Báo Cáo", use_container_width=True):
    if not supabase:
        st.error("❌ Không thể kết nối Supabase")
    else:
        try:
            # Đọc dữ liệu từ Supabase
            response = supabase.table("loi_thi").select("*").execute()
            df = pd.DataFrame(response.data) if response.data else pd.DataFrame()
            
            if df.empty:
                st.warning("⚠️ Chưa có dữ liệu để xuất.")
            else:
                # Áp dụng bộ lọc
                if filter_type == "Một ngày":
                    df = df[df['ngay'] == str(filter_date)]
                elif filter_type == "Từ ngày đến ngày":
                    df = df[(df['ngay'] >= str(filter_date_from)) & (df['ngay'] <= str(filter_date_to))]
                
                if df.empty:
                    st.warning("⚠️ Không có dữ liệu cho khoảng thời gian đã chọn.")
                else:
                    # Gom nhóm và tính tổng
                    df_tong_hop = df.groupby('ngay')[columns].sum().reset_index()
                    
                    # Tính tổng cộng
                    total_row = {'ngay': 'TỔNG CỘNG'}
                    for col in columns:
                        total_row[col] = df_tong_hop[col].sum()
                    
                    # Thêm hàng tổng cộng vào đầu
                    df_tong_hop = pd.concat([pd.DataFrame([total_row]), df_tong_hop], ignore_index=True)
                    
                    # Thêm cột STT
                    stt = [''] + list(range(1, len(df_tong_hop)))
                    df_tong_hop.insert(0, 'STT', stt)
                    
                    # Đổi tên cột
                    df_tong_hop.columns = ['STT', 'Ngày sát hạch'] + full_error_names
                    
                    # Tạo file Excel
                    excel_data = create_report_excel(df_tong_hop)
                    
                    st.download_button(
                        label="📥 Tải File Excel",
                        data=excel_data,
                        file_name=f"Bao_cao_loi_{date.today().strftime('%d_%m_%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                    # Hiển thị preview
                    with st.expander("👁️ Xem trước dữ liệu"):
                        df_display = df_tong_hop.copy()
                        df_display['Ngày sát hạch'] = df_display['Ngày sát hạch'].apply(
                            lambda x: format_date(x) if x != 'TỔNG CỘNG' else x
                        )
                        st.dataframe(df_display, use_container_width=True)
        except Exception as e:
            st.error(f"❌ Lỗi tạo báo cáo: {e}")

# ===== QUẢN LÝ DỮ LIỆU =====
with st.expander("⚙️ Quản Lý Dữ Liệu"):
    st.subheader("Xem Dữ Liệu")
    col_manage1, col_manage2 = st.columns(2)
    
    with col_manage1:
        if st.button("🔍 Xem Tất Cả Dữ Liệu", use_container_width=True):
            if supabase:
                try:
                    response = supabase.table("loi_thi").select("*").execute()
                    df_view = pd.DataFrame(response.data) if response.data else pd.DataFrame()
                    
                    if not df_view.empty:
                        # Đổi tên cột
                        df_view_display = df_view.copy()
                        df_view_display.columns = ['ID', 'Ngày sát hạch'] + full_error_names
                        
                        # Format ngày
                        df_view_display['Ngày sát hạch'] = df_view_display['Ngày sát hạch'].apply(format_date)
                        
                        # Xóa cột ID
                        df_view_display = df_view_display.drop('ID', axis=1)
                        
                        st.dataframe(df_view_display, use_container_width=True)
                    else:
                        st.info("Không có dữ liệu")
                except Exception as e:
                    st.error(f"❌ Lỗi đọc dữ liệu: {e}")
    
    with col_manage2:
        st.subheader("Xóa Dữ Liệu")
        password_input = st.text_input("Nhập mật khẩu:", type="password", key="del_password")
        if st.button("🗑️ Xóa Tất Cả", use_container_width=True):
            if password_input == "123":
                if supabase:
                    try:
                        supabase.table("loi_thi").delete().neq("id", -1).execute()
                        st.success("✅ Đã xóa tất cả dữ liệu")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Lỗi xóa dữ liệu: {e}")
            else:
                st.error("❌ Mật khẩu không chính xác!")
